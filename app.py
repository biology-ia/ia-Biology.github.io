
from flask import Flask, render_template, request, redirect, url_for
from openpyxl import load_workbook, Workbook
import os

app = Flask(__name__, static_url_path='/static')
EXCEL_FILE = 'usuarios.xlsx'

# Crear archivo Excel si no existe
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(['username', 'password'])
    wb.save(EXCEL_FILE)

def is_user_valid(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

def register_user(username, password):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return False
    ws.append([username, password])
    wb.save(EXCEL_FILE)
    return True

@app.route('/')
def home():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    error = ""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if is_user_valid(username, password):
            return redirect('/chat')  # Redirige a la página de chat
        else:
            error = "Usuario o contraseña incorrectos"
    return render_template('login.html', error=error)

@app.route('/register', methods=['GET', 'POST'])
def register():
    message = ""
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        if register_user(username, password):
            message = "Registro exitoso. Ahora puedes iniciar sesión."
        else:
            message = "El usuario ya está registrado."
    return render_template('register.html', message=message)

@app.route('/chat')
def chat():
    return '<h1>Bienvenido al chat</h1>'

if __name__ == '__main__':
    app.run(debug=True)
